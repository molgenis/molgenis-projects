__author__ = 'mslofstra'
import string
import sys
import openpyxl.styles

class ExcelSheet:
    def __init__(self, sheet):
        self.sheet = sheet

    def alter(self, column, row, value):
        """function alter_sheet
        This function can alter a sheet. It needs the sheet object to be altered,
        the cell which should be altered and the value that should be in the cell.
        """
        self.sheet[column+str(row)] = value

    def alter_header(self, column, value):
        self.alter(column, 1, value)
        return self.get_header()

    def get_column(self, index):
        col_values = []
        for row in list(self.sheet.iter_rows())[1::]:
            value = row[index].value
            col_values.append(value)
        return col_values
    def get_column_values(self, index):
        col_values = []
        for row in list(self.sheet.iter_rows())[1::]:
            value = row[index].value
            if value != None and value != 'id':
                col_values.append(value)
        return col_values

    def get_header(self, header_row=0):
        first_row = list(self.sheet.rows)[header_row]
        return [cell.value for cell in first_row if cell.value]

    def find_first_empty_column(self):
        header = self.get_header()
        return len(header)

    def write_new_column(self, column, head):
        index = self.find_first_empty_column()+1
        self.overwrite_column(column, head, index)

    def overwrite_column(self, column, head, index):
        self.sheet.cell(row=1, column=index, value=head)
        for i, item in enumerate(column):
            self.sheet.cell(row=i+2, column=index, value=item)

    def del_column(self, col_num, colmax):
        rowmax = self.find_first_empty_cell()
        for c in range(col_num, colmax):
            for r in range(1, rowmax):
                self.alter(c, r, self.sheet.cell(row=r, column=c+1).value)
        for r in range(1, rowmax):
            self.alter(colmax, r,'')
        return self.get_header()

    def find_first_empty_row(self, header_row=0):
        headers = self.get_header(header_row)
        header_letters = [letter for header, letter in zip(headers, list(string.ascii_uppercase))]
        first_empty_row = 0
        for letter in header_letters:
            first_empty_cell = self.find_first_empty_cell(letter)
            if first_empty_cell > first_empty_row:
                first_empty_row = first_empty_cell
        return first_empty_row

    def find_first_empty_cell(self, column="A"):
        """function find_first_empty_cell
        This function finds the first empty cell in the given sheet (checks if the first column is empty)
        Returns the row number of the cell."""
        iteration = 1
        empty_cell_not_found = True
        while empty_cell_not_found:
            #check if the cell is empty and contains no background color
            if self.sheet[column+str(iteration)].value== None and type(self.sheet[column+str(iteration)].fill.bgColor.indexed) != int:
                empty_cell_not_found= False
            else:
                iteration += 1
        return iteration
