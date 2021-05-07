import os
import pandas as pd
import argparse
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime


parser=argparse.ArgumentParser(description="command line args")
parser.add_argument("-out", type=str, dest='output', required=True,\
                    help="Define filename without extension to save output")

args=parser.parse_args()


def main():
    wb = Workbook()
    wb.save(args.output + "_emx.xlsx")  
    wb.close()
    
    for file_name in os.listdir():       
        if 'll_to_tgo' in file_name:
            print("Reading: " + file_name)

            template = pd.ExcelFile(file_name)
            data = template.parse()

            if "NEXT_NR, v_567" in data.columns:
                data = pd.read_excel(file_name,converters={'NEXT_NR, v_567':str})
            elif "NEXT_NR, v_280" in data.columns:
                data = pd.read_excel(file_name,converters={'NEXT_NR, v_280':str})
            elif "NEXT_NR, dupl1_v_280" in data.columns:
                data = pd.read_excel(file_name,converters={'NEXT_NR, dupl1_v_280':str})
            

            #derive entity name from file name
            entity_name = get_entity_name(file_name)

            #remove columns
            data = remove_columns(data)

            #change column names
            data = change_col_names(data)

            #change date formats
            data.loc[:, 'v_75'] = data['v_75'].apply(change_date_format)

            #remove spaces in otherwise empty fields
            data = remove_spaces(data)

##            #divide strings in notes >255 characters
##            data = divide_strings(data, entity_name)

            #write to excel in separate tabs
            write_emx(data, entity_name)

    #remove original empty wb sheet
    wb = load_workbook(args.output + "_emx.xlsx")
    wb.remove(wb['Sheet'])
    wb.save(args.output + "_emx.xlsx")
    wb.close()


def get_entity_name(s):
    #get entity_name from s

    entity_name = s[:s.find('_ll')]
    
    return entity_name


def remove_columns(df):
    col_names = df.columns
    col_idx = [idx for idx, s in enumerate(col_names) if 'LL_NR' in s][0]
    df = df.drop(df.columns[col_idx], axis=1)

    return df


def change_col_names(df):
    new_col_names = []
    
    for col_name in df.columns:
        if 'dupl' in col_name:
            start = col_name.find('dupl')
        elif 'EPDS_1' in col_name:
            start = col_name.find('EPDS_1')
        else:
            start = col_name.find('v_')
        new_col_name = col_name[start:]
        new_col_names.append(new_col_name)

    df.columns = new_col_names

    return df


def change_date_format(x):
    if x.__class__ == pd._libs.tslibs.timestamps.Timestamp:
        x = x.to_pydatetime()
    
    if x.__class__ == datetime:
        x = x.strftime("%Y-%m-%d")
    elif x.__class__ == str:
        try:
            x = datetime.strptime(x, '%Y-%b-%d').date()
            x = x.strftime("%Y-%m-%d")
        except ValueError:
            try:
                x = datetime.strptime(x, '%Y-%m-%d').date()
                x = x.strftime("%Y-%m-%d")
            except ValueError:
                x = x[:10]
                
    return x


##def divide_strings(df, entity_name):
##    for col in df:
##        new_col = col + '_continued'
##        df[new_col] = ''
##
##        i = 0
##        for item in df[col]:
##            if type(item) == str:
##                if len(item) > 255:
##                    df[col].loc[i] = item[:255]
##                    df[new_col].loc[i] = item[255:]
##                    print('String cut in: ' + entity_name + " " + col)
##            i+=1
##
##        if (df[new_col] == '').all():
##            df = df.drop(new_col, axis=1)
##
##    #and again, some answers are very long
##    for col in df:
##        new_col = col + '_2'
##        df[new_col] = ''
##
##        i = 0
##        for item in df[col]:
##            if type(item) == str:
##                if len(item) > 255:
##                    df[col].loc[i] = item[:255]
##                    df[new_col].loc[i] = item[255:]
##                    print('String cut in: ' + entity_name + " " + col)
##            i+=1
##
##        if (df[new_col] == '').all():
##            df = df.drop(new_col, axis=1) 
##            
##
##    return df


def remove_spaces(df):
    df = df.replace(" ", "")

    return df


def write_emx(df, entity_name):
    print("Writing emx of: " + entity_name)

    with pd.ExcelWriter(args.output + "_emx.xlsx", mode='a', engine='openpyxl')\
         as writer:        
        df.to_excel(writer, sheet_name = 'qs_data_' + entity_name, index=None)
        writer.save()
        writer.close()


if __name__ == '__main__':
    main()
