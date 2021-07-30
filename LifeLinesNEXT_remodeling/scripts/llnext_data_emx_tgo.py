import os
import pandas as pd
import argparse
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime


parser=argparse.ArgumentParser(description="command line args")
parser.add_argument("-out", type=str, dest='output', required=True,\
                    help="Define filename without extension to save output")

args=parser.parse_args()


def main():
    wb = Workbook()
    wb.save(args.output + "_emx.xlsx")  
    wb.close()

##    emx_attributes = pd.read_excel("C:/Users/brend/projects/LifeLinesNext/data/emx_validated/qs_emx.xlsx", sheet_name="attributes")
    emx_attributes = pd.read_excel("C:/Users/brend/projects/LifeLinesNext/data/emx_validated/ROME_emx.xlsx", sheet_name="attributes")


    for file_name in os.listdir():
        #specifics for TGO data
        if 'participants_data' in file_name or 'data_project' in file_name:
            data = pd.read_csv(file_name, sep=";")
            print('Read: ' + file_name)
            
            #derive entity name from file name
            entity_name = get_entity_name(file_name)        

            #change date formats
            data.loc[:, 'v_75'] = data['datetime']
            data.loc[:, 'v_75'] = data['v_75'].apply(change_date)

            #remove columns
            data = remove_columns(data)

            #remove rows
            data = remove_rows(data, entity_name)

            #change id format and add 0's to start
            data = change_id(data, entity_name)

            #remove None values
            data = data.replace(['-66', '-77', '-99'], ['', '', ''])
            data = data.replace([-66, -77, -99], ['', '', ''])
            data = remove_xref_zeros(data, entity_name, emx_attributes)

##            #divide strings in notes >255 characters
##            data = divide_strings(data, entity_name)

            #update data according to changes made to model
            data = update(data, entity_name)

            #write to excel in separate tabs
            write_emx(data, entity_name)

    #remove original empty wb sheet
    wb = load_workbook(args.output + "_emx.xlsx")
    wb.remove(wb['Sheet'])
    wb.save(args.output + "_emx.xlsx")
    wb.close()


def get_entity_name(s):
    #get entity_name from s
    if 'participants_data' in s:
        entity_name = s[34:s.find('_', 34)]
    elif 'data_project' in s:
        entity_name = s[13:s.find('_', 13)].replace(' ', '_')

    if entity_name == 'V1':
        entity_name = 'Father_1'
    elif entity_name == 'V2':
        entity_name = 'Father_2'
    elif entity_name == 'V3':
        entity_name = 'Father_3'
    
    return entity_name


def change_date(x):
    x = x[:10]
                            
    return x


def change_id(df, entity_name):

    if 'Mother' in entity_name or 'Father' in entity_name or 'P' in entity_name\
       or 'ROME' in entity_name:
        df = df.astype({'v_567': 'str'})

        i = 0
        for x in df['v_567']:
            if len(x) == 5:
                df.loc[i, 'v_567'] = '0' + x
            elif len(x) == 4:
                df.loc[i, 'v_567'] = '00' + x
            elif len(x) == 3:
                df.loc[i, 'v_567'] = '000' + x
            elif len(x) == 2:
                df.loc[i, 'v_567'] = '0000' + x
            elif len(x) == 1:
                df.loc[i, 'v_567'] = '00000' + x
            i+=1

    elif 'Baby' in entity_name:
        if entity_name in ['M1_Baby', 'M2_Baby', 'W2_Baby']:
            df = df.astype({'v_280': 'str'})

            i=0
            for x in df['v_280']:
                if len(x) == 5:
                    df.loc[i, 'v_280'] = '0' + x
                elif len(x) == 4:
                    df.loc[i, 'v_280'] = '00' + x
                elif len(x) == 3:
                    df.loc[i, 'v_280'] = '000' + x
                elif len(x) == 2:
                    df.loc[i, 'v_280'] = '0000' + x
                elif len(x) == 1:
                    df.loc[i, 'v_280'] = '00000' + x
                i+=1
        else:
            df = df.astype({'dupl1_v_280': 'str'})

            i=0
            for x in df['dupl1_v_280']:
                if len(x) == 5:
                    df.loc[i, 'dupl1_v_280'] = '0' + x
                elif len(x) == 4:
                    df.loc[i, 'dupl1_v_280'] = '00' + x
                elif len(x) == 3:
                    df.loc[i, 'dupl1_v_280'] = '000' + x
                elif len(x) == 2:
                    df.loc[i, 'dupl1_v_280'] = '0000' + x
                elif len(x) == 1:
                    df.loc[i, 'dupl1_v_280'] = '00000' + x
                i+=1    
          
    return df


def remove_columns(df):
    col_names = df.columns
    col_idx = [idx for idx, value in enumerate(col_names) if 'v_' not in value]
    df = df.drop(df.columns[col_idx], axis=1)

    return df


def remove_rows(df, entity_name):
    #remove rows when id column is empty

    if 'Mother' in entity_name or 'Father' in entity_name or 'P' in entity_name:
        row_idx = [idx for idx, value in enumerate(df['v_567']) if '-66' in value]
        df = df.drop(row_idx, axis=0)

    elif 'Baby' in entity_name:
        if entity_name in ['M1_Baby', 'M2_Baby', 'W2_Baby']:
            row_idx = [idx for idx, value in enumerate(df['v_280']) if '-66' in value]
            df = df.drop(row_idx, axis=0)

        else:
            row_idx = [idx for idx, value in enumerate(df['dupl1_v_280']) if '-66' in value]
            df = df.drop(row_idx, axis=0)

    return df

def remove_xref_zeros(df, entity_name, emx_attributes):
    #remove zeros from xrefs, they encode missing values, but are not in TGO codebooks
    for column in df:
        if column in emx_attributes['name'].to_list():
            #get column index in emx_attributes['name'] for entity_name
            i = 0
            for name in emx_attributes['name']:
                if name == column and entity_name in emx_attributes['entity'][i] \
                   and emx_attributes['dataType'][i] == 'xref':
                    if not name == 'v_279':
                        df.loc[:, column] = df[column].apply(remove_zeros)
                i+=1
    return df



def remove_zeros(x):
    if x == 0:
        return ''
    else:
        return x
		
    

##def divide_strings(df, entity_name):
##    for col in df:
##        new_col = col + '_continued'
##        df[new_col] = ''
##
##        i = 0
##        for item in df[col]:
##            if type(item) == str:
##                if len(item) > 255:
##                    df.loc[i, col] = item[:255]
##                    df.loc[i, new_col] = item[255:]
##                    print('String cut in: ' + entity_name + " " + col + +str(len(item)))
##            i+=1
##
##        if (df[new_col] == '').all():
##            df = df.drop(new_col, axis=1) 
##
##    return df


def update(df, entity_name):
    if entity_name in ['P12', 'P32']:
        df = update_P12_P32(df)
    if entity_name == 'P18':
        df = update_P18(df)
    if entity_name == 'Father_2':
        df = update_Father_2(df)
    if entity_name in ['M2_Baby', 'M9_Baby']:
        df = update_M2_M9_Baby(df)
    if entity_name in ['M2_Baby', 'M6_Baby', 'M9_Baby']:
        df = update_M2_M6_M9_Baby(df)

    return df


def update_P12_P32(df):
    df.rename(columns={'v_468':'v_521',
                       'v_470':'v_522',
                       'v_472':'v_523',
                       'v_474':'v_524',
                       'v_476':'v_525',
                       'v_478':'v_526'}, inplace=True)

    if 'v_467' in df.columns:
        df = df.assign(v_671 = '')
        df = df.assign(v_672 = '')
        df = df.assign(v_673 = '')
        df = df.assign(v_674 = '')
        df = df.assign(v_675 = '')
        df = df.assign(v_676 = '')
        df = df.assign(v_677 = '')
        df = df.assign(v_678 = '')
        df = df.assign(v_679 = '')

        df.loc[df['v_467'] == 1, 'v_671'] = 1
        df.loc[df['v_467'] == 2, 'v_672'] = 1
        df.loc[df['v_467'] == 3, 'v_673'] = 1
        df.loc[df['v_467'] == 4, 'v_674'] = 1
        df.loc[df['v_467'] == 5, 'v_675'] = 1
        df.loc[df['v_467'] == 6, 'v_676'] = 1
        df.loc[df['v_467'] == 7, 'v_677'] = 1
        df.loc[df['v_467'] == 8, 'v_678'] = 1
        df.loc[df['v_467'] == 9, 'v_679'] = 1
                    
        df = df.drop('v_467', axis=1)

    if 'v_469' in df.columns:
        df = df.assign(v_626 = '')
        df = df.assign(v_627 = '')
        df = df.assign(v_628 = '')
        df = df.assign(v_629 = '')
        df = df.assign(v_630 = '')
        df = df.assign(v_631 = '')
        df = df.assign(v_632 = '')
        df = df.assign(v_633 = '')
        df = df.assign(v_634 = '')

        df.loc[df['v_469'] == 1, 'v_626'] = 1
        df.loc[df['v_469'] == 2, 'v_627'] = 1
        df.loc[df['v_469'] == 3, 'v_628'] = 1
        df.loc[df['v_469'] == 4, 'v_629'] = 1
        df.loc[df['v_469'] == 5, 'v_630'] = 1
        df.loc[df['v_469'] == 6, 'v_631'] = 1
        df.loc[df['v_469'] == 7, 'v_632'] = 1
        df.loc[df['v_469'] == 8, 'v_633'] = 1
        df.loc[df['v_469'] == 9, 'v_634'] = 1
        
        df = df.drop('v_469', axis=1)

    if 'v_471' in df.columns:
        df = df.assign(v_635 = '')
        df = df.assign(v_636 = '')
        df = df.assign(v_637 = '')
        df = df.assign(v_638 = '')
        df = df.assign(v_639 = '')
        df = df.assign(v_640 = '')
        df = df.assign(v_641 = '')
        df = df.assign(v_642 = '')
        df = df.assign(v_643 = '')

        df.loc[df['v_471'] == 1, 'v_635'] = 1
        df.loc[df['v_471'] == 2, 'v_636'] = 1
        df.loc[df['v_471'] == 3, 'v_637'] = 1
        df.loc[df['v_471'] == 4, 'v_638'] = 1
        df.loc[df['v_471'] == 5, 'v_639'] = 1
        df.loc[df['v_471'] == 6, 'v_640'] = 1
        df.loc[df['v_471'] == 7, 'v_641'] = 1
        df.loc[df['v_471'] == 8, 'v_642'] = 1
        df.loc[df['v_471'] == 9, 'v_643'] = 1
        
        df = df.drop('v_471', axis=1)

    if 'v_473' in df.columns:
        df = df.assign(v_644 = '')
        df = df.assign(v_645 = '')
        df = df.assign(v_646 = '')
        df = df.assign(v_647 = '')
        df = df.assign(v_648 = '')
        df = df.assign(v_649 = '')
        df = df.assign(v_650 = '')
        df = df.assign(v_651 = '')
        df = df.assign(v_652 = '')

        df.loc[df['v_473'] == 1, 'v_644'] = 1
        df.loc[df['v_473'] == 2, 'v_645'] = 1
        df.loc[df['v_473'] == 1, 'v_646'] = 1
        df.loc[df['v_473'] == 4, 'v_647'] = 1
        df.loc[df['v_473'] == 5, 'v_648'] = 1
        df.loc[df['v_473'] == 6, 'v_649'] = 1
        df.loc[df['v_473'] == 7, 'v_650'] = 1
        df.loc[df['v_473'] == 8, 'v_651'] = 1
        df.loc[df['v_473'] == 9, 'v_652'] = 1

        df = df.drop('v_473', axis=1)

    if 'v_475' in df.columns:
        df = df.assign(v_653 = '')
        df = df.assign(v_654 = '')
        df = df.assign(v_655 = '')
        df = df.assign(v_656 = '')
        df = df.assign(v_657 = '')
        df = df.assign(v_658 = '')
        df = df.assign(v_659 = '')
        df = df.assign(v_660 = '')
        df = df.assign(v_661 = '')

        df.loc[df['v_475'] == 1, 'v_653'] = 1
        df.loc[df['v_475'] == 2, 'v_654'] = 1
        df.loc[df['v_475'] == 3, 'v_655'] = 1
        df.loc[df['v_475'] == 4, 'v_656'] = 1
        df.loc[df['v_475'] == 5, 'v_657'] = 1
        df.loc[df['v_475'] == 6, 'v_658'] = 1
        df.loc[df['v_475'] == 7, 'v_659'] = 1
        df.loc[df['v_475'] == 8, 'v_660'] = 1
        df.loc[df['v_475'] == 9, 'v_661'] = 1
        
        df = df.drop('v_475', axis=1)

    if 'v_477' in df.columns:
        df = df.assign(v_662 = '')
        df = df.assign(v_663 = '')
        df = df.assign(v_664 = '')
        df = df.assign(v_665 = '')
        df = df.assign(v_666 = '')
        df = df.assign(v_667 = '')
        df = df.assign(v_668 = '')
        df = df.assign(v_669 = '')
        df = df.assign(v_670 = '')

        df.loc[df['v_477'] == 1, 'v_662'] = 1
        df.loc[df['v_477'] == 2, 'v_663'] = 1
        df.loc[df['v_477'] == 3, 'v_664'] = 1
        df.loc[df['v_477'] == 4, 'v_665'] = 1
        df.loc[df['v_477'] == 5, 'v_666'] = 1
        df.loc[df['v_477'] == 6, 'v_667'] = 1
        df.loc[df['v_477'] == 7, 'v_668'] = 1
        df.loc[df['v_477'] == 8, 'v_669'] = 1
        df.loc[df['v_477'] == 9, 'v_670'] = 1
        
        df = df.drop('v_477', axis=1)

    return df


def update_P18(df):
    if 'v_429' in df.columns:
        df = df.assign(v_506 = '')
        df = df.assign(v_507 = '')
        df = df.assign(v_508 = '')
        df = df.assign(v_509 = '')
        df = df.assign(v_510 = '')
        df = df.assign(v_511 = '')
        df = df.assign(v_512 = '')
        df = df.assign(v_513 = '')
        
        df.loc[df['v_429'] == 1, 'v_506'] = 1
        df.loc[df['v_429'] == 2, 'v_507'] = 1
        df.loc[df['v_429'] == 3, 'v_508'] = 1
        df.loc[df['v_429'] == 4, 'v_509'] = 1
        df.loc[df['v_429'] == 5, 'v_510'] = 1
        df.loc[df['v_429'] == 6, 'v_511'] = 1
        df.loc[df['v_429'] == 7, 'v_512'] = 1
        df.loc[df['v_429'] == 8, 'v_513'] = 1

        df = df.drop('v_429', axis=1)
        
    if 'v_432' in df.columns:
        df = df.assign(v_521 = '')
        df = df.assign(v_522 = '')
        df = df.assign(v_523 = '')
        df = df.assign(v_524 = '')
        df = df.assign(v_525 = '')
        df = df.assign(v_526 = '')
        df = df.assign(v_527 = '')
        df = df.assign(v_528 = '')
        df = df.assign(v_529 = '')
        
        df.loc[df['v_432'] == 1, 'v_521'] = 1
        df.loc[df['v_432'] == 2, 'v_522'] = 1
        df.loc[df['v_432'] == 3, 'v_523'] = 1
        df.loc[df['v_432'] == 4, 'v_524'] = 1
        df.loc[df['v_432'] == 5, 'v_525'] = 1
        df.loc[df['v_432'] == 6, 'v_526'] = 1
        df.loc[df['v_432'] == 7, 'v_527'] = 1
        df.loc[df['v_432'] == 8, 'v_528'] = 1
        df.loc[df['v_432'] == 9, 'v_529'] = 1

        df = df.drop('v_432', axis=1)

    if not 'v_514' in df.columns and 'v_431' in df.columns:
        df.rename(columns={'v_431':'temp'}, inplace=True)

        df = df.assign(v_431 = '')
        df = df.assign(v_514 = '')
        df = df.assign(v_515 = '')
        df = df.assign(v_516 = '')
        df = df.assign(v_517 = '')
        df = df.assign(v_518 = '')
        df = df.assign(v_519 = '')
        df = df.assign(v_520 = '')

        df.loc[df['v_506'] == 1, 'v_431'] = df['temp']  
        df.loc[df['v_507'] == 1, 'v_514'] = df['temp']
        df.loc[df['v_508'] == 1, 'v_515'] = df['temp']
        df.loc[df['v_509'] == 1, 'v_516'] = df['temp']
        df.loc[df['v_510'] == 1, 'v_517'] = df['temp']
        df.loc[df['v_511'] == 1, 'v_518'] = df['temp']
        df.loc[df['v_512'] == 1, 'v_519'] = df['temp']
        df.loc[df['v_513'] == 1, 'v_520'] = df['temp']

        df = df.drop('temp', axis=1)
        
    if 'v_434' in df.columns:       
        df = df.assign(v_431 = '')
        df = df.assign(v_514 = '')
        df = df.assign(v_515 = '')
        df = df.assign(v_516 = '')
        df = df.assign(v_517 = '')
        df = df.assign(v_518 = '')
        df = df.assign(v_519 = '')
        df = df.assign(v_520 = '')

        df.loc[df['v_523'] == 1, 'v_530'] = df['v_434']
        df.loc[df['v_524'] == 1, 'v_531'] = df['v_434']
        df.loc[df['v_525'] == 1, 'v_532'] = df['v_434']
        df.loc[df['v_526'] == 1, 'v_533'] = df['v_434']
        df.loc[df['v_527'] == 1, 'v_534'] = df['v_434']
        df.loc[df['v_528'] == 1, 'v_535'] = df['v_434']
        df.loc[df['v_529'] == 1, 'v_536'] = df['v_434']

        df = df.drop('v_434', axis=1)             
        
    return df

def update_Father_2(df):
    if 'v_107' in df.columns and not 'v_339' in df.columns:
        df.rename(columns={'v_107':'temp'}, inplace=True)
        
        df = df.assign(v_107 = '')
        df = df.assign(v_339 = '')
        df = df.assign(v_350 = '')
        df = df.assign(v_361 = '')
        df = df.assign(v_372 = '')
        
        df.loc[df['temp'] == 1, 'v_107'] = 1
        df.loc[df['temp'] == 2, 'v_339'] = 1
        df.loc[df['temp'] == 3, 'v_350'] = 1
        df.loc[df['temp'] == 4, 'v_361'] = 1
        df.loc[df['temp'] == 5, 'v_372'] = 1
        
        df = df.drop('temp', axis=1)

    if 'v_108' in df.columns and not 'v_340' in df.columns:
        df.rename(columns={'v_108':'temp'}, inplace=True)
        
        df = df.assign(v_108 = '')
        df = df.assign(v_340 = '')
        df = df.assign(v_351 = '')
        df = df.assign(v_362 = '')
        df = df.assign(v_373 = '')
        
        df.loc[df['temp'] == 1, 'v_108'] = 1
        df.loc[df['temp'] == 2, 'v_340'] = 1
        df.loc[df['temp'] == 3, 'v_351'] = 1
        df.loc[df['temp'] == 4, 'v_362'] = 1
        df.loc[df['temp'] == 5, 'v_373'] = 1
        
        df = df.drop('temp', axis=1)

    if 'v_109' in df.columns and not 'v_341' in df.columns:
        df.rename(columns={'v_109':'temp'}, inplace=True)
        
        df = df.assign(v_109 = '')
        df = df.assign(v_341 = '')
        df = df.assign(v_352 = '')
        df = df.assign(v_363 = '')
        df = df.assign(v_374 = '')
        
        df.loc[df['temp'] == 1, 'v_109'] = 1
        df.loc[df['temp'] == 2, 'v_341'] = 1
        df.loc[df['temp'] == 3, 'v_352'] = 1
        df.loc[df['temp'] == 4, 'v_363'] = 1
        df.loc[df['temp'] == 5, 'v_374'] = 1
        
        df = df.drop('temp', axis=1)

    if 'v_110' in df.columns and not 'v_342' in df.columns:
        df.rename(columns={'v_110':'temp'}, inplace=True)
        
        df = df.assign(v_110 = '')
        df = df.assign(v_342 = '')
        df = df.assign(v_353 = '')
        df = df.assign(v_364 = '')
        df = df.assign(v_375 = '')
        
        df.loc[df['temp'] == 1, 'v_110'] = 1
        df.loc[df['temp'] == 2, 'v_342'] = 1
        df.loc[df['temp'] == 3, 'v_353'] = 1
        df.loc[df['temp'] == 4, 'v_364'] = 1
        df.loc[df['temp'] == 5, 'v_375'] = 1
        
        df = df.drop('temp', axis=1)

    if 'v_111' in df.columns and not 'v_343' in df.columns:
        df.rename(columns={'v_111':'temp'}, inplace=True)
        
        df = df.assign(v_111 = '')
        df = df.assign(v_343 = '')
        df = df.assign(v_354 = '')
        df = df.assign(v_365 = '')
        df = df.assign(v_376 = '')
        
        df.loc[df['temp'] == 1, 'v_111'] = 1
        df.loc[df['temp'] == 2, 'v_343'] = 1
        df.loc[df['temp'] == 3, 'v_354'] = 1
        df.loc[df['temp'] == 4, 'v_365'] = 1
        df.loc[df['temp'] == 5, 'v_376'] = 1
        
        df = df.drop('temp', axis=1)

    if 'v_113' in df.columns and not 'v_344' in df.columns:
        df.rename(columns={'v_113':'temp'}, inplace=True)
        
        df = df.assign(v_113 = '')
        df = df.assign(v_344 = '')
        df = df.assign(v_355 = '')
        df = df.assign(v_366 = '')
        df = df.assign(v_377 = '')
        
        df.loc[df['temp'] == 1, 'v_113'] = 1
        df.loc[df['temp'] == 2, 'v_344'] = 1
        df.loc[df['temp'] == 3, 'v_355'] = 1
        df.loc[df['temp'] == 4, 'v_366'] = 1
        df.loc[df['temp'] == 5, 'v_377'] = 1
        
        df = df.drop('temp', axis=1)

    if 'v_114' in df.columns and not 'v_345' in df.columns:
        df.rename(columns={'v_114':'temp'}, inplace=True)
        
        df = df.assign(v_114 = '')
        df = df.assign(v_345 = '')
        df = df.assign(v_356 = '')
        df = df.assign(v_367 = '')
        df = df.assign(v_378 = '')
        
        df.loc[df['temp'] == 1, 'v_114'] = 1
        df.loc[df['temp'] == 2, 'v_345'] = 1
        df.loc[df['temp'] == 3, 'v_356'] = 1
        df.loc[df['temp'] == 4, 'v_367'] = 1
        df.loc[df['temp'] == 5, 'v_378'] = 1
        
        df = df.drop('temp', axis=1)

    if 'v_115' in df.columns and not 'v_346' in df.columns:
        df.rename(columns={'v_115':'temp'}, inplace=True)
        
        df = df.assign(v_115 = '')
        df = df.assign(v_346 = '')
        df = df.assign(v_357 = '')
        df = df.assign(v_368 = '')
        df = df.assign(v_379 = '')
        
        df.loc[df['temp'] == 1, 'v_115'] = 1
        df.loc[df['temp'] == 2, 'v_346'] = 1
        df.loc[df['temp'] == 3, 'v_357'] = 1
        df.loc[df['temp'] == 4, 'v_368'] = 1
        df.loc[df['temp'] == 5, 'v_379'] = 1
        
        df = df.drop('temp', axis=1)

    if 'v_116' in df.columns and not 'v_347' in df.columns:
        df.rename(columns={'v_116':'temp'}, inplace=True)
        
        df = df.assign(v_116 = '')
        df = df.assign(v_347 = '')
        df = df.assign(v_358 = '')
        df = df.assign(v_369 = '')
        df = df.assign(v_380 = '')
        
        df.loc[df['temp'] == 1, 'v_116'] = 1
        df.loc[df['temp'] == 2, 'v_347'] = 1
        df.loc[df['temp'] == 3, 'v_358'] = 1
        df.loc[df['temp'] == 4, 'v_369'] = 1
        df.loc[df['temp'] == 5, 'v_380'] = 1
        
        df = df.drop('temp', axis=1)

    if 'v_117' in df.columns and not 'v_348' in df.columns:
        df.rename(columns={'v_117':'temp'}, inplace=True)
        
        df = df.assign(v_117 = '')
        df = df.assign(v_348 = '')
        df = df.assign(v_359 = '')
        df = df.assign(v_370 = '')
        df = df.assign(v_381 = '')
        
        df.loc[df['temp'] == 1, 'v_117'] = 1
        df.loc[df['temp'] == 2, 'v_348'] = 1
        df.loc[df['temp'] == 3, 'v_359'] = 1
        df.loc[df['temp'] == 4, 'v_370'] = 1
        df.loc[df['temp'] == 5, 'v_381'] = 1
        
        df = df.drop('temp', axis=1)

    if 'v_118' in df.columns and not 'v_349' in df.columns:
        df.rename(columns={'v_118':'temp'}, inplace=True)
        
        df = df.assign(v_118 = '')
        df = df.assign(v_349 = '')
        df = df.assign(v_360 = '')
        df = df.assign(v_371 = '')
        df = df.assign(v_382 = '')
        
        df.loc[df['temp'] == 1, 'v_118'] = 1
        df.loc[df['temp'] == 2, 'v_349'] = 1
        df.loc[df['temp'] == 3, 'v_360'] = 1
        df.loc[df['temp'] == 4, 'v_371'] = 1
        df.loc[df['temp'] == 5, 'v_382'] = 1
        
        df = df.drop('temp', axis=1)

    return df

def update_M2_M9_Baby(df):
    df.rename(columns={'v_189':'temp'}, inplace=True)
    df = df.assign(v_189 = '')

    df.loc[df['temp'] == 1, 'v_189'] = 3
    df.loc[df['temp'] == 2, 'v_189'] = 3
    df.loc[df['temp'] == 3, 'v_189'] = 3
    df.loc[df['temp'] == 4, 'v_189'] = 4
    df.loc[df['temp'] == 5, 'v_189'] = 5

    df = df.drop('temp', axis=1)

    return df

def update_M2_M6_M9_Baby(df):
    df.rename(columns={'v_195':'temp'}, inplace=True)
    df = df.assign(v_195 = '')

    df.loc[df['temp'] == 1, 'v_195'] = 3
    df.loc[df['temp'] == 2, 'v_195'] = 3
    df.loc[df['temp'] == 3, 'v_195'] = 3
    df.loc[df['temp'] == 4, 'v_195'] = 4
    df.loc[df['temp'] == 5, 'v_195'] = 5

    df = df.drop('temp', axis=1)

    return df
    

def write_emx(df, entity_name):
    print("Writing emx of: " + entity_name)

    with pd.ExcelWriter(args.output + "_emx.xlsx", mode='a', engine='openpyxl')\
         as writer:        
        df.to_excel(writer, sheet_name = 'qs_data_' + entity_name, index=None)
        writer.save()
        writer.close()


if __name__ == '__main__':
    main()
