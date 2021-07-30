import xlrd
import pandas as pd
import argparse
import re
from openpyxl import Workbook
from openpyxl import load_workbook
import os


parser=argparse.ArgumentParser(description="command line args")
parser.add_argument("-out", type=str, dest='output', required=True,\
                    help="Define filename without extension to save output")

args=parser.parse_args()


def main():
    all_tgo_df = pd.DataFrame()
    entity_list = []
    
    for file_name in os.listdir():
        if file_name.startswith('codebook'):
            #read tgo_next codebook
            c_sh = read_sheet(file_name)

            #derive entity name from codebook file name
            entity_name = get_entity(file_name)
            #save to entity_list
            entity_list.append(entity_name)
            print(entity_list)

            #read variable name, data type, question, coding and
            #description and store in df
            tgo_df = extract_data(c_sh, entity_name)           

            #match to ll next variable name using combined codebook
            cc_file_name = get_file_name(entity_name)

            if not cc_file_name == None:
                combi_codebook = read_table(cc_file_name)
            else:
                combi_codebook = None
            tgo_df = get_ll_name_and_compound(combi_codebook, cc_file_name, tgo_df, entity_name)

            #combine into one df, with variable entity ('Father1', 'Father2', etc.)
            all_tgo_df = pd.concat([all_tgo_df, tgo_df], ignore_index=True)
           
    #define molgenis attributes
    all_tgo_df, codes = define_attributes(all_tgo_df)

    print(len(codes))

    #write emx to file
    write_emx(all_tgo_df, codes, entity_list)
    
        #create entities from coding.
            # check for existing coding first, assign xref
            # otherwise create new entity 

    #save all_tgo_df in Excel for data checks
    all_tgo_df.to_excel(args.output + '.xlsx', index=None)

    
def read_sheet(excel_file):
    #open workbook by name and sheet by index
    wb = xlrd.open_workbook(excel_file)
    sheet = wb.sheet_by_index(0)
    print('Read TGO NEXT codebook: ' + excel_file)

    return sheet


def get_entity(s):
    #get entity_name from s
    entity_name = s[17:s.find('_', 17)]
    
    return entity_name
    

def extract_data(sh, entity_name):
    #extract data model from codebook
    df = pd.DataFrame(columns=['tgo_name', 'tgo_type', 'tgo_question', \
                               'tgo_coding', 'description', 'entity'])


    df = extract_model(df, sh, entity_name)
    
    return df


def get_file_name(entity_name):
    #get file name of combined codebook using entity_name
    file_name = None
    
    for file in os.listdir():
        if entity_name in file and 'output' in file:
            file_name = file
            break
    
    return file_name


def read_table(file_name):
    print('Read combined codebook: ' + file_name)
    #read combined codebook into pandas df
    template = pd.ExcelFile(file_name)
    df = template.parse()

    return df


def get_ll_name_and_compound(combi_codebook, cc_file_name, tgo_df, entity_name):
    #find index j of tgo_name in combi_codebook and save associated
    #ll_name and compound in tgo_df at index i
    tgo_df['ll_name'] = ''
    tgo_df['compound'] = ''

    if not cc_file_name == None:    
        i = 0
        for name in tgo_df['tgo_name']:
            if name in combi_codebook.tgo_name.to_list():
                j = combi_codebook.tgo_name[combi_codebook.tgo_name == name].index[0]
                if not combi_codebook['ll_name'][j] == 'not_present':
                    ll_name = combi_codebook['ll_name'][j]
                    if '_' in ll_name and not 'LL_NR' in ll_name:
                        tgo_df['compound'][i], tgo_df['ll_name'][i] = \
                                               split_string(ll_name)
                    else:
                        tgo_df['ll_name'][i] = ll_name
            else:
                print('Not in combi_codebook: ' + entity_name, str(i), name)
    ##            print(name)
                
            i+=1

    return tgo_df           


def split_string(s):
    start = s[:s.find('_')]
    end = s[s.find('_') + 1:]              

    return start, end
    


def extract_model(df, tgo_sh, entity):
    #extract all possible questions and associated variables
    #from first column of tgo sheet and return in df
    idx = 0
    
    for i in range(14, tgo_sh.nrows - 1):
        s = tgo_sh.cell_value(i, 0)
        if not s == '' and not s.startswith('v_') and \
           not s == ' ' and not type_float(s):

            tgo_name = tgo_sh.cell_value(i+1, 1)
            tgo_type = tgo_sh.cell_value(i+1, 2)
            tgo_question = clean_string(s)
            tgo_coding = get_coding(tgo_sh, tgo_type, i)
            if not tgo_question == '':
                description = tgo_question + '; ' + tgo_sh.cell_value(i+1, 3)
            else:
                description = tgo_sh.cell_value(i+1, 3)

            #store in df
            df.loc[idx] = [tgo_name, tgo_type, tgo_question, tgo_coding,\
                           description, entity]

            idx += 1

        elif s.startswith('v_') and (tgo_sh.cell_value(i-1, 0) == ''\
           or tgo_sh.cell_value(i-1, 0) == ' ' or tgo_sh.cell_value(i-1, 0).startswith('v_')):
            tgo_name = tgo_sh.cell_value(i, 1)
            tgo_type = tgo_sh.cell_value(i, 2)
            tgo_coding = get_coding(tgo_sh, tgo_type, i-1)
            if not tgo_question == '':
                description = tgo_question + '; ' + tgo_sh.cell_value(i, 3)
            else:
                description = tgo_sh.cell_value(i, 3)
                
            #store in df
            df.loc[idx] = [tgo_name, tgo_type, tgo_question, tgo_coding,\
                           description, entity]
                                                   
            idx += 1

    return df


def define_attributes(df):
    df['entity'] = df['entity'].apply(clean_entity)
    df['name'] = df['tgo_name']
    df['label'] = df['ll_name'].apply(get_label)
    df['dataType'] = df['tgo_type'].apply(get_data_type)
    df['idAttribute'] = df['tgo_name'].apply(get_id_attribute)
    df['partOfAttribute'] = df['compound']
    df, codes = get_ref_entity(df)

    return df, codes


def clean_entity(x):
    x = x.replace(' ', '_')

    return x


def get_label(x):
    if x == 'ID_LL_NR':
        return 'NEXT_NR'
    else:
        return x
    

def get_data_type(x):    
    if x == 'int':
        return 'xref'
    else:
        return 'string'


def get_id_attribute(x):
    if x == 'v_567':
        return 'TRUE'
    else:
        return ''
    

def get_ref_entity(df):
    coding_list = []
    codes = {}
    df['refEntity'] = ''
        
    for idx, value in df['tgo_coding'].iteritems():
        ref_entity = 'codes_' + df['tgo_name'][idx]

        if df['dataType'][idx] == 'xref':
            if value in coding_list:
                df['refEntity'][idx] = list(codes.keys())[list(codes.values()\
                                                               ).index(value)]
                
            else:
                coding_list.append(value)
                if ref_entity in codes:
                    ref_entity = ref_entity + '_2'
                if ref_entity in codes:
                    ref_entity = ref_entity + '_3'
                if ref_entity in codes:
                    ref_entity = ref_entity + '_4'
##                print(ref_entity, df['entity'][idx])
                codes[ref_entity] = value
##                print(len(codes), len(coding_list))
                df['refEntity'][idx] = ref_entity
                
    return df, codes


def write_emx(df, codes, entity_list):
    print("Writing emx")
    wb = Workbook()
    wb.save(args.output + "_emx.xlsx")

    file_name = args.output + "_emx.xlsx"

    book = load_workbook(file_name)
    writer = pd.ExcelWriter(file_name, engine = 'openpyxl')
    writer.book = book

    #write different components of emx
    write_packages(writer)
    write_entities(codes, entity_list, writer)
    write_attributes(df, codes, writer)
    write_coding(codes, writer)
    writer.save()
    writer.close()

    #remove original empty wb sheet
    wb = load_workbook(args.output + "_emx.xlsx")
    wb.remove(wb['Sheet'])
    wb.save(args.output + "_emx.xlsx")
    wb.close()

def write_packages(writer):

    row_1 = ['codes', 'Coding names and values', '']
    row_2 = ['qs_data', 'Questionnaire data', '']
    
    packages_df = pd.DataFrame([row_1, row_2],\
                               columns = ['name', 'label', 'description'])
    
    packages_df.to_excel(writer, sheet_name = 'packages', index=None)


def write_entities(codes, entity_list, writer):
    #write entities to tab in excel file  
    entities_df = pd.DataFrame(columns = ['name', 'package', 'label',\
                                          'description'])
    idx = 0
    #save enitities from entity_list to entities_df
    for entity in entity_list:
        name = entity.replace(' ', '_')
        package = 'qs_data'
        label = ''
        description = ''
        entities_df.loc[idx] = [name, package, label, description]
        idx += 1
        
    #save entity names from codes dictionary to entities_df
    for key in codes:
        name = key
        package = 'codes'
        label = ''
        description = ''
        entities_df.loc[idx] = [name, package, label, description]
        idx += 1  

    #write to tab in excel file
    entities_df.to_excel(writer, sheet_name = 'entities', index=None)
                               

def write_attributes(df, codes, writer):
    #define columns from df containing all questionnaire data and defined columns
    #for emx
    attributes_df = df[['entity', 'name', 'label', 'dataType', 'refEntity', \
                 'description', 'idAttribute', 'partOfAttribute']]

    attributes_df.loc[:, 'entity'] = 'qs_data_' + attributes_df['entity']

    #define attributes for xref coding from codes dictionary   
    for key in codes:
        row_1 = [key, 'id', 'Id', 'int', '', '', 'TRUE', '']
        row_2 = [key, 'label', 'Label', 'string', '', '', '', '']

        xref_attributes = pd.DataFrame([row_1, row_2],\
                                       columns = ['entity', 'name', 'label', \
                                                  'dataType', 'refEntity', \
                                                  'description', 'idAttribute',\
                                                  'partOfAttribute'])

        #append each defined xref to attributes_df
        attributes_df = attributes_df.append(xref_attributes)

    #define compounds
    attributes_df = define_compounds(attributes_df)
    
    #write to tab in excel file
    attributes_df.to_excel(writer, sheet_name = 'attributes', index=None)


def define_compounds(df):
    df2 = df
    compound_list = []
    
    i = 0
    j = 0

    col_idx = df.columns.get_loc('partOfAttribute')
    
    for attribute in df['partOfAttribute']:
        if not i == 0 and not attribute == '':
            if not df.iloc[i, col_idx] == df.iloc[i-1, col_idx]:
                
                compound_line = pd.DataFrame({'entity': df['entity'][i], \
                                              'name': df['partOfAttribute'][i].lower(), \
                                              'label': df['partOfAttribute'][i], \
                                              'dataType': 'compound', \
                                              'refEntity': '', \
                                              'description': '', \
                                              'idAttribute': '', \
                                            'partOfAttribute': ''}, index=[i+1])

                compound = df['entity'][i] + df['partOfAttribute'][i]
                if compound not in compound_list:
                    df2 = pd.concat([df2.iloc[:j], compound_line, \
                                 df.iloc[i:]]).reset_index(drop=True)
                    compound_list.append(compound)
                    j+=1                    
                
        i+=1
        j+=1

    #attributes references in lower case
    df2.loc[:, 'partOfAttribute'] = df2['partOfAttribute'].apply(lower)

    return df2


def lower(x):
    x = x.lower()

    return x



def write_coding(codes, writer):
    #write coding to tables in tabs in excel file
    i = 0
    for key in codes:#is xref_entity_dict, not coding list
        series = pd.Series(codes[key], name='label')
        series.index.name = 'id'
        df = series.reset_index()
        df.to_excel(writer, sheet_name = key, index=None)
        i+=1
  

def get_coding(tgo_sh, tgo_type, i):
    #get coding and values
    coding = ''

    if tgo_type == 'int':
        coding = {}
        for j in range(1, 30):
            if not tgo_sh.cell_value(i+1+j, 2) in ['', 'varchar',\
                                                   'blob', 'int']:
                label = int(tgo_sh.cell_value(i+1+j, 2))
                value = tgo_sh.cell_value(i+1+j, 3)
##                code = label + ' = ' + value                               
                coding[label] = value
            else:
                break

    return coding


def clean_string(s):
    #remove non-breaking spaces
    s = s.replace(u"\u00A0", " ")
    
    #remove text between brackets and brackets themselves
    if '(' in s:
        s = s[:s.find('(')] + s[s.find(')')+1:]

    if '(' in s:
        s = s[:s.find('(')] + s[s.find(')')+1:]

    #remove leading and trailing whitespace
    s = s.strip()

    return s


def type_float(string):
    try:
        float(string[0])
        return True
    except ValueError:
        return False


if __name__ == '__main__':
    main()
